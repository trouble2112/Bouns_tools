"""创建带完整Excel公式的模板"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
wb.remove(wb.active)

# 样式
HEADER_FILL = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
INPUT_FILL = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')
CALC_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
RESULT_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# ========== 1. 参数表 ==========
ws_param = wb.create_sheet('参数设置')
params = [
    ['参数名称', '参数值', '说明'],
    ['1月时间系数', 1.15, ''],
    ['2月时间系数', 1.15, ''],
    ['3月时间系数', 1.10, ''],
    ['4月时间系数', 1.00, ''],
    ['5月时间系数', 0.90, ''],
    ['6月时间系数', 0.85, ''],
    ['90%回款门槛', 0.85, '回款率>=此值才发90%奖'],
    ['100%回款门槛', 0.90, '回款率>=此值才发100%奖'],
    ['常委固定补贴', 60000, '半年总额'],
    ['销售月补贴', 800, '新购/高校每月'],
    ['DM叠加模式', 'exclusive', 'exclusive=只发最高档, stack=叠加'],
    ['其他叠加模式', 'stack', 'exclusive=只发最高档, stack=叠加'],
]
for r, row in enumerate(params, 1):
    for c, val in enumerate(row, 1):
        cell = ws_param.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        if r == 1:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        elif c == 2 and r > 1:
            cell.fill = INPUT_FILL
ws_param.column_dimensions['A'].width = 15
ws_param.column_dimensions['B'].width = 12
ws_param.column_dimensions['C'].width = 30

# ========== 2. 人员数据表 ==========
ws_data = wb.create_sheet('人员数据')
headers = ['序号','姓名','岗位','区域','组织单元',
           '1月产值','2月产值','3月产值','4月产值','5月产值','6月产值',
           '分公司产值','年度目标','回款率',
           '区域90%','区域100%','全国90%','全国100%',
           '分配比例','CEO奖金']

for c, h in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER

# 预填50行输入区
for r in range(2, 52):
    for c in range(1, 21):
        cell = ws_data.cell(row=r, column=c)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER

# 数据验证 - 岗位下拉
role_dv = DataValidation(type='list', formula1='"CP,DM,VP,MGR,SALES_USER,SALES_NEW,SALES_EDU"')
ws_data.add_data_validation(role_dv)
role_dv.add('C2:C51')

# 数据验证 - 是否下拉
yesno_dv = DataValidation(type='list', formula1='"是,否"')
ws_data.add_data_validation(yesno_dv)
yesno_dv.add('O2:R51')

for c in range(1, 21):
    ws_data.column_dimensions[get_column_letter(c)].width = 10 if c < 5 else 9

# ========== 3. 计算结果表(带公式) ==========
ws_result = wb.create_sheet('计算结果')
result_headers = ['序号','姓名','岗位','产值合计','完成率',
                  '过程激励','完成奖90%','完成奖100%','完成奖小计',
                  '区域奖','全国奖','固定补贴','CEO奖金','奖金合计']

for c, h in enumerate(result_headers, 1):
    cell = ws_result.cell(row=1, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER

# 写入公式 (行2-51)
for r in range(2, 52):
    row_str = str(r)
    
    # A: 序号
    ws_result.cell(row=r, column=1).value = f'=IF(人员数据!B{row_str}<>"",人员数据!A{row_str},"")'
    ws_result.cell(row=r, column=1).border = THIN_BORDER
    
    # B: 姓名
    ws_result.cell(row=r, column=2).value = f'=IF(人员数据!B{row_str}<>"",人员数据!B{row_str},"")'
    ws_result.cell(row=r, column=2).border = THIN_BORDER
    
    # C: 岗位
    ws_result.cell(row=r, column=3).value = f'=IF(人员数据!B{row_str}<>"",人员数据!C{row_str},"")'
    ws_result.cell(row=r, column=3).border = THIN_BORDER
    
    # D: 产值合计
    cell = ws_result.cell(row=r, column=4)
    cell.value = f'=IF(人员数据!B{row_str}<>"",SUM(人员数据!F{row_str}:K{row_str}),"")'
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # E: 完成率 (分公司产值/年度目标)
    cell = ws_result.cell(row=r, column=5)
    cell.value = f'=IF(AND(人员数据!B{row_str}<>"",人员数据!M{row_str}>0),人员数据!L{row_str}/人员数据!M{row_str},"")'
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '0.0%'
    
    # F: 过程激励
    # CP=0, DM/VP=0.4%, MGR=1%, SALES_USER=2%, SALES_NEW/EDU=3%
    incentive = f'''=IF(人员数据!B{row_str}="","",
IF(人员数据!C{row_str}="CP",0,
IF(OR(人员数据!C{row_str}="DM",人员数据!C{row_str}="VP"),
(人员数据!F{row_str}*参数设置!$B$2+人员数据!G{row_str}*参数设置!$B$3+人员数据!H{row_str}*参数设置!$B$4+人员数据!I{row_str}*参数设置!$B$5+人员数据!J{row_str}*参数设置!$B$6+人员数据!K{row_str}*参数设置!$B$7)*0.004,
IF(人员数据!C{row_str}="MGR",
(人员数据!F{row_str}*参数设置!$B$2+人员数据!G{row_str}*参数设置!$B$3+人员数据!H{row_str}*参数设置!$B$4+人员数据!I{row_str}*参数设置!$B$5+人员数据!J{row_str}*参数设置!$B$6+人员数据!K{row_str}*参数设置!$B$7)*0.01,
IF(人员数据!C{row_str}="SALES_USER",
(人员数据!F{row_str}*参数设置!$B$2+人员数据!G{row_str}*参数设置!$B$3+人员数据!H{row_str}*参数设置!$B$4+人员数据!I{row_str}*参数设置!$B$5+人员数据!J{row_str}*参数设置!$B$6+人员数据!K{row_str}*参数设置!$B$7)*0.02,
(人员数据!F{row_str}*参数设置!$B$2+人员数据!G{row_str}*参数设置!$B$3+人员数据!H{row_str}*参数设置!$B$4+人员数据!I{row_str}*参数设置!$B$5+人员数据!J{row_str}*参数设置!$B$6+人员数据!K{row_str}*参数设置!$B$7)*0.03)))))'''
    cell = ws_result.cell(row=r, column=6)
    cell.value = incentive.replace('\n', '')
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # G: 完成奖90%
    bonus90 = f'''=IF(人员数据!B{row_str}="","",
IF(人员数据!C{row_str}="CP",0,
IF(AND(E{row_str}>=0.9,人员数据!N{row_str}>=参数设置!$B$8),
IF(人员数据!C{row_str}="DM",MIN(人员数据!L{row_str}*0.004,40000),
人员数据!L{row_str}*0.015*IF(人员数据!S{row_str}<>"",人员数据!S{row_str},1)),0)))'''
    cell = ws_result.cell(row=r, column=7)
    cell.value = bonus90.replace('\n', '')
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # H: 完成奖100%
    bonus100 = f'''=IF(人员数据!B{row_str}="","",
IF(人员数据!C{row_str}="CP",0,
IF(AND(E{row_str}>=1,人员数据!N{row_str}>=参数设置!$B$9),
IF(人员数据!C{row_str}="DM",MIN(人员数据!L{row_str}*0.004,40000),
人员数据!L{row_str}*0.015*IF(人员数据!S{row_str}<>"",人员数据!S{row_str},1)),0)))'''
    cell = ws_result.cell(row=r, column=8)
    cell.value = bonus100.replace('\n', '')
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # I: 完成奖小计 (根据叠加模式)
    bonus_total = f'''=IF(人员数据!B{row_str}="","",
IF(人员数据!C{row_str}="DM",
IF(参数设置!$B$12="exclusive",MAX(G{row_str},H{row_str}),G{row_str}+H{row_str}),
IF(参数设置!$B$13="exclusive",MAX(G{row_str},H{row_str}),G{row_str}+H{row_str})))'''
    cell = ws_result.cell(row=r, column=9)
    cell.value = bonus_total.replace('\n', '')
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # J: 区域奖
    region = f'''=IF(人员数据!B{row_str}="","",
IF(人员数据!C{row_str}="CP",IF(人员数据!O{row_str}="是",30000,0)+IF(人员数据!P{row_str}="是",30000,0),
IF(人员数据!C{row_str}="DM",IF(OR(人员数据!O{row_str}="是",人员数据!P{row_str}="是"),40000,0),0)))'''
    cell = ws_result.cell(row=r, column=10)
    cell.value = region.replace('\n', '')
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # K: 全国奖 (仅CP)
    national = f'''=IF(人员数据!B{row_str}="","",
IF(人员数据!C{row_str}="CP",IF(人员数据!Q{row_str}="是",40000,0)+IF(人员数据!R{row_str}="是",40000,0),0))'''
    cell = ws_result.cell(row=r, column=11)
    cell.value = national.replace('\n', '')
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # L: 固定补贴
    subsidy = f'''=IF(人员数据!B{row_str}="","",
IF(人员数据!C{row_str}="CP",参数设置!$B$10,
IF(OR(人员数据!C{row_str}="SALES_NEW",人员数据!C{row_str}="SALES_EDU"),参数设置!$B$11*6,0)))'''
    cell = ws_result.cell(row=r, column=12)
    cell.value = subsidy.replace('\n', '')
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # M: CEO奖金
    cell = ws_result.cell(row=r, column=13)
    cell.value = f'=IF(人员数据!B{row_str}<>"",IF(人员数据!T{row_str}<>"",人员数据!T{row_str},0),"")'
    cell.border = THIN_BORDER
    cell.fill = CALC_FILL
    cell.number_format = '#,##0'
    
    # N: 奖金合计
    cell = ws_result.cell(row=r, column=14)
    cell.value = f'=IF(人员数据!B{row_str}<>"",F{row_str}+I{row_str}+J{row_str}+K{row_str}+L{row_str}+M{row_str},"")'
    cell.border = THIN_BORDER
    cell.fill = RESULT_FILL
    cell.number_format = '#,##0'
    cell.font = Font(bold=True)

# 设置列宽
widths = [6,10,12,12,10,12,12,12,12,10,10,10,10,12]
for c, w in enumerate(widths, 1):
    ws_result.column_dimensions[get_column_letter(c)].width = w

# ========== 4. 汇总表 ==========
ws_sum = wb.create_sheet('汇总')
ws_sum['A1'] = '2026上半年奖金汇总'
ws_sum['A1'].font = Font(size=16, bold=True)

ws_sum['A3'] = '总人数:'
ws_sum['B3'] = '=COUNTA(计算结果!B2:B51)'
ws_sum['B3'].font = Font(bold=True)

ws_sum['A4'] = '奖金总额:'
ws_sum['B4'] = '=SUM(计算结果!N2:N51)'
ws_sum['B4'].number_format = '¥#,##0'
ws_sum['B4'].font = Font(bold=True, size=14, color='1976D2')

ws_sum['A6'] = '按岗位汇总'
ws_sum['A6'].font = Font(bold=True)
ws_sum['A7'] = '常委CP'
ws_sum['B7'] = '=SUMIF(计算结果!C:C,"CP",计算结果!N:N)'
ws_sum['B7'].number_format = '#,##0'
ws_sum['A8'] = '总经理DM'
ws_sum['B8'] = '=SUMIF(计算结果!C:C,"DM",计算结果!N:N)'
ws_sum['B8'].number_format = '#,##0'
ws_sum['A9'] = '副总经理VP'
ws_sum['B9'] = '=SUMIF(计算结果!C:C,"VP",计算结果!N:N)'
ws_sum['B9'].number_format = '#,##0'
ws_sum['A10'] = '部门经理MGR'
ws_sum['B10'] = '=SUMIF(计算结果!C:C,"MGR",计算结果!N:N)'
ws_sum['B10'].number_format = '#,##0'
ws_sum['A11'] = '销售(用户)'
ws_sum['B11'] = '=SUMIF(计算结果!C:C,"SALES_USER",计算结果!N:N)'
ws_sum['B11'].number_format = '#,##0'
ws_sum['A12'] = '销售(新购)'
ws_sum['B12'] = '=SUMIF(计算结果!C:C,"SALES_NEW",计算结果!N:N)'
ws_sum['B12'].number_format = '#,##0'
ws_sum['A13'] = '销售(高校)'
ws_sum['B13'] = '=SUMIF(计算结果!C:C,"SALES_EDU",计算结果!N:N)'
ws_sum['B13'].number_format = '#,##0'

ws_sum.column_dimensions['A'].width = 15
ws_sum.column_dimensions['B'].width = 15

# ========== 5. 填入测试数据 ==========
test_data = [
    [1, '王总', 'CP', '全国', '总部', 0,0,0,0,0,0, 0, 0, 0.95, '是','是','是','否', '', 50000],
    [2, '李总', 'DM', '华北', '北京分公司', 500000,600000,700000,800000,750000,650000, 4000000, 3800000, 0.92, '是','否','否','否', '', 20000],
    [3, '张经理', 'MGR', '华东', '上海分公司', 150000,180000,200000,220000,190000,160000, 3000000, 2800000, 0.91, '否','否','否','否', 0.25, 5000],
    [4, '陈销售', 'SALES_NEW', '华东', '上海分公司', 80000,90000,100000,110000,95000,85000, 3000000, 500000, 0.88, '否','否','否','否', 0.15, 0],
    [5, '刘副总', 'VP', '华南', '深圳分公司', 200000,250000,280000,300000,270000,230000, 2500000, 2300000, 0.93, '否','否','否','否', 0.4, 10000],
    [6, '赵销售', 'SALES_EDU', '西南', '成都分公司', 60000,70000,85000,90000,80000,65000, 1800000, 400000, 0.90, '否','否','否','否', 0.2, 0],
]

for row_idx, row_data in enumerate(test_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_data.cell(row=row_idx, column=col_idx, value=value)

# 保存
output_path = r'c:\Users\EDWD\PG2\bonus_calculator\bonus_with_formulas.xlsx'
wb.save(output_path)
print(f'带公式的Excel已生成: {output_path}')
print('包含6条测试数据，在腾讯文档打开后可直接查看计算结果')
