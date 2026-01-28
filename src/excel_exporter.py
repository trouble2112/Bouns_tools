"""
2026上半年奖金计算引擎 - Excel导出模块
Export calculation results to Excel
"""
from typing import List, Tuple
from models import PersonData, BonusDetail, ValidationResult
from bonus_engine import BonusCalculator
from config import GlobalConfig, Role

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """Excel报表导出器"""
    
    def __init__(self, config: GlobalConfig = None):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        # 样式定义（在__init__中初始化，确保openpyxl已导入）
        self.HEADER_FILL = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        self.HEADER_FONT = Font(color="FFFFFF", bold=True)
        self.INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        self.CALC_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        self.WARNING_FILL = PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid")
        self.ERROR_FILL = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
        self.THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.config = config or GlobalConfig()
    
    def create_template(self, filepath: str):
        """创建空白Excel模板"""
        wb = openpyxl.Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建各工作表
        self._create_cover_sheet(wb)
        self._create_params_sheet(wb)
        self._create_data_sheet(wb)
        self._create_result_sheet(wb)
        self._create_validation_sheet(wb)
        self._create_dict_sheet(wb)
        
        wb.save(filepath)
        print(f"模板已创建: {filepath}")
    
    def export_results(
        self, 
        results: List[Tuple[BonusDetail, ValidationResult]],
        filepath: str
    ):
        """导出计算结果到Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 创建汇总表
        self._create_summary_sheet(wb, results)
        
        # 创建明细表
        self._create_detail_sheet(wb, results)
        
        wb.save(filepath)
        print(f"结果已导出: {filepath}")
    
    def _create_cover_sheet(self, wb):
        """创建首页"""
        ws = wb.create_sheet("首页")
        
        ws.merge_cells('B2:F2')
        ws['B2'] = "2026上半年奖金计算工具 V1.0"
        ws['B2'].font = Font(size=20, bold=True)
        
        info = [
            ("版本", "1.0"),
            ("更新日期", "2026-01-28"),
            ("适用范围", "2026年1-6月奖金计算"),
        ]
        
        for i, (label, value) in enumerate(info, start=4):
            ws[f'B{i}'] = label
            ws[f'C{i}'] = value
        
        ws['B7'] = "使用步骤："
        steps = [
            "1. 确认「全局参数」表中的参数设置",
            "2. 在「人员数据」表录入人员信息和产值",
            "3. 查看「计算结果」表的奖金明细",
            "4. 检查「数据校验」表确认无异常"
        ]
        for i, step in enumerate(steps, start=8):
            ws[f'B{i}'] = step
        
        ws['B13'] = "⚠️ 待业务确认项目（见「全局参数」表黄色标记）"
        ws['B13'].fill = self.WARNING_FILL
        
        ws.column_dimensions['B'].width = 50
    
    def _create_params_sheet(self, wb):
        """创建参数表"""
        ws = wb.create_sheet("全局参数")
        
        headers = ["参数名称", "参数代码", "默认值", "当前值", "说明", "待确认"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        params = [
            ("1月时间系数", "TIME_COEFF_01", 1.15, 1.15, "1月产值激励系数", ""),
            ("2月时间系数", "TIME_COEFF_02", 1.15, 1.15, "2月产值激励系数", ""),
            ("3月时间系数", "TIME_COEFF_03", 1.10, 1.10, "3月产值激励系数", ""),
            ("4月时间系数", "TIME_COEFF_04", 1.00, 1.00, "4月产值激励系数", ""),
            ("5月时间系数", "TIME_COEFF_05", 0.90, 0.90, "5月产值激励系数", ""),
            ("6月时间系数", "TIME_COEFF_06", 0.85, 0.85, "6月产值激励系数", ""),
            ("", "", "", "", "", ""),
            ("90%奖回款门槛", "THRESHOLD_90", 0.85, 0.85, "发放90%完成奖的最低回款率", "✓"),
            ("100%奖回款门槛", "THRESHOLD_100", 0.90, 0.90, "发放100%完成奖的最低回款率", "✓"),
            ("", "", "", "", "", ""),
            ("DM叠加模式", "DM_STACK_MODE", "exclusive", "exclusive", "exclusive=仅发最高档; stack=累计", "✓"),
            ("其他岗位叠加模式", "OTHER_STACK_MODE", "stack", "stack", "副总/部门经理/销售的叠加模式", "✓"),
            ("", "", "", "", "", ""),
            ("完成率计算模式", "COMPLETION_MODE", "from_target", "from_target", "from_target=自动计算; manual=手填", ""),
            ("是否拆分发放显示", "SHOW_PAYOUT_SPLIT", "FALSE", "FALSE", "是否显示50%即时/50%回款后", "✓"),
            ("", "", "", "", "", ""),
            ("常委固定补贴(半年)", "CP_SUBSIDY", 60000, 60000, "常委固定补贴总额", ""),
            ("销售月补贴", "SALES_MONTHLY_SUB", 800, 800, "新购/高校销售月度补贴", ""),
        ]
        
        for row_idx, row_data in enumerate(params, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                
                # 当前值列可编辑
                if col_idx == 4 and value != "":
                    cell.fill = self.INPUT_FILL
                
                # 待确认标记
                if col_idx == 6 and value == "✓":
                    cell.fill = self.WARNING_FILL
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 8
    
    def _create_data_sheet(self, wb):
        """创建数据录入表"""
        ws = wb.create_sheet("人员数据")
        
        headers = [
            "序号", "姓名", "岗位", "区域", "组织单元",
            "1月产值", "2月产值", "3月产值", "4月产值", "5月产值", "6月产值",
            "分公司总产值", "年度目标", "完成率(手填)", "回款率",
            "区域完成90%", "区域完成100%", "全国完成90%", "全国完成100%",
            "个人分配比例", "CEO奖金",
            "产值合计", "完成率(计算)", "数据状态"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
        
        # 设置输入区样式（前21列）
        for row in range(2, 52):  # 预留50行
            for col in range(1, 22):
                cell = ws.cell(row=row, column=col)
                cell.fill = self.INPUT_FILL
                cell.border = self.THIN_BORDER
            
            # 计算列（22-24列）
            for col in range(22, 25):
                cell = ws.cell(row=row, column=col)
                cell.fill = self.CALC_FILL
                cell.border = self.THIN_BORDER
        
        # 添加数据验证
        # 岗位下拉
        role_validation = DataValidation(
            type="list",
            formula1='"CP,DM,VP,MGR,SALES_USER,SALES_NEW,SALES_EDU"',
            allow_blank=False
        )
        role_validation.error = "请选择有效的岗位代码"
        ws.add_data_validation(role_validation)
        role_validation.add('C2:C51')
        
        # 区域下拉
        region_validation = DataValidation(
            type="list",
            formula1='"华北,华东,华南,西南,西北,东北,全国"',
            allow_blank=False
        )
        ws.add_data_validation(region_validation)
        region_validation.add('D2:D51')
        
        # 是/否下拉
        yesno_validation = DataValidation(
            type="list",
            formula1='"是,否"',
            allow_blank=False
        )
        ws.add_data_validation(yesno_validation)
        yesno_validation.add('P2:S51')
        
        # 设置列宽
        for col in range(1, 25):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _create_result_sheet(self, wb):
        """创建计算结果表"""
        ws = wb.create_sheet("计算结果")
        
        headers = [
            "序号", "姓名", "岗位", "区域",
            "1月激励", "2月激励", "3月激励", "4月激励", "5月激励", "6月激励", "过程激励小计",
            "完成奖90%", "完成奖100%", "完成奖小计",
            "区域奖90%", "区域奖100%", "区域奖小计",
            "全国奖90%", "全国奖100%", "全国奖小计",
            "固定补贴", "CEO奖金",
            "奖金合计", "叠加模式", "备注"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
        
        # 设置列宽
        for col in range(1, 26):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _create_validation_sheet(self, wb):
        """创建数据校验表"""
        ws = wb.create_sheet("数据校验")
        
        ws['A1'] = "数据校验结果"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = "总记录数:"
        ws['B3'] = "=COUNTA(人员数据!B:B)-1"
        
        ws['A4'] = "错误数量:"
        ws['B4'] = "=COUNTIF(人员数据!X:X,\"异常\")"
        ws['B4'].fill = self.ERROR_FILL
        
        ws['A6'] = "异常明细"
        ws['A6'].font = Font(bold=True)
        
        headers = ["行号", "姓名", "问题描述", "严重程度"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
    
    def _create_dict_sheet(self, wb):
        """创建字典表"""
        ws = wb.create_sheet("字典表")
        
        # 岗位映射
        ws['A1'] = "岗位代码"
        ws['B1'] = "岗位名称"
        ws['C1'] = "过程激励比例"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        ws['C1'].font = Font(bold=True)
        
        roles = [
            ("CP", "常委", 0),
            ("DM", "总经理", 0.004),
            ("VP", "副总经理", 0.004),
            ("MGR", "部门经理", 0.01),
            ("SALES_USER", "销售-用户部", 0.02),
            ("SALES_NEW", "销售-新购", 0.03),
            ("SALES_EDU", "销售-高校", 0.03),
        ]
        
        for row, (code, name, rate) in enumerate(roles, start=2):
            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=rate)
        
        # 区域列表
        ws['E1'] = "区域代码"
        ws['F1'] = "区域名称"
        ws['E1'].font = Font(bold=True)
        ws['F1'].font = Font(bold=True)
        
        regions = [
            ("NORTH", "华北"),
            ("EAST", "华东"),
            ("SOUTH", "华南"),
            ("SOUTHWEST", "西南"),
            ("NORTHWEST", "西北"),
            ("NORTHEAST", "东北"),
        ]
        
        for row, (code, name) in enumerate(regions, start=2):
            ws.cell(row=row, column=5, value=code)
            ws.cell(row=row, column=6, value=name)
    
    def _create_summary_sheet(self, wb, results: List[Tuple[BonusDetail, ValidationResult]]):
        """创建汇总表"""
        ws = wb.create_sheet("汇总报表")
        
        ws['A1'] = "2026上半年奖金汇总报表"
        ws['A1'].font = Font(size=16, bold=True)
        
        # 按岗位汇总
        ws['A3'] = "按岗位汇总"
        ws['A3'].font = Font(bold=True)
        
        headers = ["岗位", "人数", "过程激励", "完成奖", "区域奖", "全国奖", "补贴", "CEO奖", "合计"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        # 按岗位统计
        role_stats = {}
        for detail, _ in results:
            role = detail.role.value
            if role not in role_stats:
                role_stats[role] = {
                    'count': 0,
                    'incentive': 0,
                    'completion': 0,
                    'region': 0,
                    'national': 0,
                    'subsidy': 0,
                    'ceo': 0,
                    'total': 0
                }
            stats = role_stats[role]
            stats['count'] += 1
            stats['incentive'] += detail.incentive_total
            stats['completion'] += detail.completion_bonus_total
            stats['region'] += detail.region_bonus_total
            stats['national'] += detail.national_bonus_total
            stats['subsidy'] += detail.fixed_subsidy
            stats['ceo'] += detail.ceo_bonus
            stats['total'] += detail.grand_total
        
        row = 5
        grand_total = {'count': 0, 'incentive': 0, 'completion': 0, 'region': 0, 
                       'national': 0, 'subsidy': 0, 'ceo': 0, 'total': 0}
        
        for role, stats in role_stats.items():
            ws.cell(row=row, column=1, value=role)
            ws.cell(row=row, column=2, value=stats['count'])
            ws.cell(row=row, column=3, value=stats['incentive'])
            ws.cell(row=row, column=4, value=stats['completion'])
            ws.cell(row=row, column=5, value=stats['region'])
            ws.cell(row=row, column=6, value=stats['national'])
            ws.cell(row=row, column=7, value=stats['subsidy'])
            ws.cell(row=row, column=8, value=stats['ceo'])
            ws.cell(row=row, column=9, value=stats['total'])
            
            for key in grand_total:
                grand_total[key] += stats[key]
            row += 1
        
        # 合计行
        ws.cell(row=row, column=1, value="合计")
        ws.cell(row=row, column=1).font = Font(bold=True)
        for col, key in enumerate(['count', 'incentive', 'completion', 'region', 
                                   'national', 'subsidy', 'ceo', 'total'], start=2):
            cell = ws.cell(row=row, column=col, value=grand_total[key])
            cell.font = Font(bold=True)
        
        # 格式化金额
        for r in range(5, row + 1):
            for c in range(3, 10):
                cell = ws.cell(row=r, column=c)
                cell.number_format = '#,##0'
    
    def _create_detail_sheet(self, wb, results: List[Tuple[BonusDetail, ValidationResult]]):
        """创建明细表"""
        ws = wb.create_sheet("明细")
        
        headers = [
            "序号", "姓名", "岗位", "区域", "组织单元",
            "1月激励", "2月激励", "3月激励", "4月激励", "5月激励", "6月激励", "过程激励小计",
            "完成奖90%", "完成奖100%", "完成奖小计",
            "区域奖", "全国奖", "固定补贴", "CEO奖金",
            "奖金合计", "完成率", "回款率", "叠加模式", "待确认项"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        for row_idx, (detail, validation) in enumerate(results, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=detail.name)
            ws.cell(row=row_idx, column=3, value=detail.role.value)
            ws.cell(row=row_idx, column=4, value=detail.region)
            ws.cell(row=row_idx, column=5, value=detail.org_unit)
            
            # 月度激励
            for month in range(1, 7):
                ws.cell(row=row_idx, column=5 + month, 
                       value=detail.monthly_incentives.get(month, 0))
            ws.cell(row=row_idx, column=12, value=detail.incentive_total)
            
            # 完成奖
            ws.cell(row=row_idx, column=13, value=detail.completion_bonus_90)
            ws.cell(row=row_idx, column=14, value=detail.completion_bonus_100)
            ws.cell(row=row_idx, column=15, value=detail.completion_bonus_total)
            
            # 其他奖项
            ws.cell(row=row_idx, column=16, value=detail.region_bonus_total)
            ws.cell(row=row_idx, column=17, value=detail.national_bonus_total)
            ws.cell(row=row_idx, column=18, value=detail.fixed_subsidy)
            ws.cell(row=row_idx, column=19, value=detail.ceo_bonus)
            
            # 合计
            ws.cell(row=row_idx, column=20, value=detail.grand_total)
            
            # 参考数据
            ws.cell(row=row_idx, column=21, value=f"{detail.completion_rate*100:.1f}%")
            ws.cell(row=row_idx, column=22, value=f"{detail.collection_rate*100:.1f}%")
            ws.cell(row=row_idx, column=23, value=detail.completion_bonus_mode)
            
            # 待确认项
            if detail.pending_confirmations:
                ws.cell(row=row_idx, column=24, value="; ".join(detail.pending_confirmations))
                ws.cell(row=row_idx, column=24).fill = self.WARNING_FILL
        
        # 格式化金额列
        for row in range(2, len(results) + 2):
            for col in [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]:
                ws.cell(row=row, column=col).number_format = '#,##0'


def create_excel_template(filepath: str = "bonus_calculator_template.xlsx"):
    """便捷函数：创建Excel模板"""
    exporter = ExcelExporter()
    exporter.create_template(filepath)


def export_to_excel(
    results: List[Tuple[BonusDetail, ValidationResult]],
    filepath: str = "bonus_results.xlsx"
):
    """便捷函数：导出计算结果"""
    exporter = ExcelExporter()
    exporter.export_results(results, filepath)
